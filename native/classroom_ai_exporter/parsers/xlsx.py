from __future__ import annotations

from pathlib import Path


def extract_xlsx_text(path: str | Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    sections: list[str] = []
    for sheet in workbook.worksheets:
        sections.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                sections.append(" | ".join(values))
    return "\n".join(sections)
